"""
债券批量分析工具
优化后的低耦合高内聚架构
"""

import akshare as ak
import pandas as pd
import requests
import os
import numpy as np
from datetime import datetime, time as dt_time
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import time
import re
import threading
import random
from dataclasses import dataclass, field
from typing import Optional, Dict, Any, List
import warnings


# ==================== 配置模块 ====================

@dataclass
class Config:
    """全局配置类"""
    ONLINE_MODE: bool = True
    CACHE_DIR: str = "cache"
    CACHE_FILE_BASE: str = "bond_metadata_cache"
    OUTPUT_FILE_BASE: str = "bond_analysis_results"
    CONCURRENT_THREADS: int = 1
    SAVE_INTERVAL: int = 10
    RETRY_COUNT: int = 5
    DELAY_BETWEEN_REQUESTS: float = 5.0
    
    USER_AGENTS: List[str] = field(default_factory=lambda: [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17.3 Safari/605.1.15",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36"
    ])

    # 列映射配置
    HEADER_MAPPING: Dict[str, str] = field(default_factory=lambda: {
        '债券简称': '债券简称', '债券类型': '债券类型', '剩余天数': '剩余天数',
        '剩余期限_格式化': '剩余期限', '税后年收益率': '税后年收益率',
        '到期日': '到期日', '票面利率': '票面利率', '付息频率': '付息频率',
        '付息方式': '付息方式', '加权收益率': '加权收益率',
        '最新收益率': '最新收益率', '成交净价': '成交净价',
        '交易量': '成交额(亿元)', '成交时间': '成交时间'
    })

    COLS_ORDER: List[str] = field(default_factory=lambda: [
        '债券简称', '债券类型', '剩余天数', '剩余期限_格式化', '税后年收益率',
        '到期日', '票面利率', '付息频率', '付息方式', '加权收益率',
        '最新收益率', '成交净价', '交易量', '成交时间'
    ])


config = Config()


# ==================== 缓存管理模块 ====================

class CacheManager:
    """缓存管理器 - 负责所有缓存操作"""
    
    def __init__(self, config: Config):
        self._config = config
        self._lock = threading.Lock()
    
    @property
    def cache_dir(self) -> str:
        return self._config.CACHE_DIR
    
    @property
    def cache_file_base(self) -> str:
        return self._config.CACHE_FILE_BASE
    
    def is_within_cache_window(self) -> bool:
        """判断当前时间是否处于交易日的 8:00-20:00 之间"""
        if not self._config.ONLINE_MODE:
            return False
        
        now = datetime.now()
        if now.weekday() >= 5:
            return False
        
        current_time = now.time()
        return dt_time(8, 0) <= current_time <= dt_time(20, 0)
    
    def get_latest_cache_date(self) -> tuple:
        """获取最新的缓存日期"""
        if not os.path.exists(self._config.CACHE_DIR):
            return None, None
        
        date_pattern = re.compile(r"^(\d{4}-\d{2}-\d{2})$")
        date_dirs = []
        
        print(os.listdir(self._config.CACHE_DIR))
        
        for f in os.listdir(self._config.CACHE_DIR):
            full_path = os.path.join(self._config.CACHE_DIR, f)
            if os.path.isdir(full_path):
                if date_pattern.match(f):
                    deal_cache = os.path.join(full_path, "bond_deal_cache.csv")
                    if os.path.exists(deal_cache):
                        date_dirs.append(f)
        
        if not date_dirs:
            return None, None
        
        date_dirs.sort(reverse=True)
        return date_dirs[0], date_dirs[0]
    
    def save_metadata_cache(self, cache_dict: Dict, settlement_dt_str: str) -> None:
        """安全保存元数据缓存"""
        if not cache_dict:
            return
        
        date_cache_dir = os.path.join(self._config.CACHE_DIR, settlement_dt_str)
        os.makedirs(date_cache_dir, exist_ok=True)
        cache_file = os.path.join(date_cache_dir, f"{self._config.CACHE_FILE_BASE}.csv")
        
        with self._lock:
            try:
                data_list = []
                for symbol, meta in cache_dict.items():
                    if 'symbol' not in meta:
                        meta['symbol'] = symbol
                    data_list.append(meta)
                
                df = pd.DataFrame(data_list)
                if not df.empty:
                    df.to_csv(cache_file, index=False, encoding='utf-8-sig')
            except Exception as e:
                print(f"保存缓存失败: {e}")
    
    def load_metadata_cache(self, settlement_dt_str: str) -> Dict:
        """加载元数据缓存"""
        cache = {}
        date_cache_dir = os.path.join(self._config.CACHE_DIR, settlement_dt_str)
        metadata_cache_file = os.path.join(date_cache_dir, f"{self._config.CACHE_FILE_BASE}.csv")
        
        if os.path.exists(metadata_cache_file):
            try:
                cache_df = pd.read_csv(metadata_cache_file, encoding='utf-8-sig')
                if not cache_df.empty and 'symbol' in cache_df.columns:
                    cache_df = cache_df.dropna(subset=['symbol'])
                    cache_df = cache_df.drop_duplicates(subset=['symbol'], keep='last')
                    cache = cache_df.set_index('symbol', drop=False).to_dict('index')
            except Exception as e:
                print(f"加载缓存失败 ({e})")
        
        return cache
    
    def get_deal_cache_path(self, date_str: str) -> str:
        """获取成交缓存文件路径"""
        return os.path.join(self._config.CACHE_DIR, date_str, "bond_deal_cache.csv")


# ==================== 数据获取模块 ====================

class BondDataFetcher:
    """债券数据获取器 - 负责所有API调用"""
    
    def __init__(self, config: Config):
        self._config = config
        self._session: Optional[requests.Session] = None
    
    def _create_session(self) -> requests.Session:
        """创建并配置请求会话"""
        if self._session is None:
            self._session = requests.Session()
            try:
                self._session.get("https://www.chinamoney.com.cn/chinese/zqjc/", timeout=15)
            except:
                pass
        return self._session
    
    def fetch_deal_data(self, settlement_dt_str: str, cache_manager: CacheManager) -> Optional[pd.DataFrame]:
        """获取成交数据 - 优先使用缓存"""
        deal_cache_file = cache_manager.get_deal_cache_path(settlement_dt_str)
        
        if os.path.exists(deal_cache_file):
            try:
                deal_df = pd.read_csv(deal_cache_file, encoding='utf-8-sig')
                print(f"从缓存加载 {len(deal_df)} 条成交记录。")
                return deal_df
            except Exception as e:
                print(f"加载成交行情缓存失败: {e}")
        
        if self._config.ONLINE_MODE:
            try:
                deal_df = ak.bond_spot_deal()
                print(f"获取 {len(deal_df)} 条成交记录。")
                
                date_cache_dir = os.path.join(self._config.CACHE_DIR, settlement_dt_str)
                os.makedirs(date_cache_dir, exist_ok=True)
                deal_df.to_csv(deal_cache_file, index=False, encoding='utf-8-sig')
                return deal_df
            except Exception as e:
                print(f"获取行情失败: {e}")
        
        return None
    
    def fetch_metadata(self, symbol: str, session: Optional[requests.Session] = None) -> Optional[Dict]:
        """获取单个债券元数据"""
        for attempt in range(self._config.RETRY_COUNT):
            try:
                metadata = self._fetch_metadata_impl(symbol, attempt, session)
                if metadata:
                    return metadata
            except Exception as e:
                tqdm.write(f"异常: {symbol} 抓取错误: {e}")
                time.sleep(5)
        
        return None
    
    def _fetch_metadata_impl(self, symbol: str, attempt: int, session: Optional[requests.Session] = None) -> Optional[Dict]:
        """获取债券元数据实现"""
        search_symbol = symbol.replace(" ", "")
        caller = session if session else requests
        
        # 搜索接口
        search_url = "https://www.chinamoney.com.cn/ags/ms/cm-u-bond-md/BondMarketInfoList2"
        search_payload = {
            "pageNo": "1", "pageSize": "15", "bondName": search_symbol,
            "bondCode": "", "issueEnty": "", "bondType": "", "bondSpclPrjctVrty": "",
            "couponType": "", "issueYear": "", "entyDefinedCode": "", "rtngShrt": ""
        }
        
        headers = self._build_headers()
        time.sleep(self._config.DELAY_BETWEEN_REQUESTS + random.uniform(2.0, 5.0) * (attempt + 1))
        
        r_search = caller.post(search_url, data=search_payload, headers=headers, timeout=20)
        
        if r_search.status_code in [403, 421]:
            wait_time = 30 * (attempt + 1)
            tqdm.write(f"警告: {symbol} 触发访问限制 ({r_search.status_code})，等待 {wait_time} 秒...")
            time.sleep(wait_time)
            return None
        
        if r_search.status_code != 200:
            return None
        
        search_json = r_search.json()
        result_list = search_json.get('data', {}).get('resultList', [])
        if not result_list:
            return None
        
        query_code = None
        for res in result_list:
            if res.get('bondName').replace(" ", "") == search_symbol:
                query_code = res.get('bondDefinedCode')
                break
        
        if not query_code:
            query_code = result_list[0].get('bondDefinedCode')
        
        # 详情接口
        time.sleep(random.uniform(1.5, 3.0))
        detail_url = "https://www.chinamoney.com.cn/ags/ms/cm-u-bond-md/BondDetailInfo"
        detail_headers = headers.copy()
        detail_headers["Referer"] = f"https://www.chinamoney.com.cn/chinese/zqjc/?bondDefinedCode={query_code}"
        
        r_detail = caller.post(detail_url, data={"bondDefinedCode": query_code}, headers=detail_headers, timeout=20)
        
        if r_detail.status_code != 200:
            return None
        
        res_json = r_detail.json()
        if 'data' not in res_json or 'bondBaseInfo' not in res_json['data']:
            return None
        
        data = res_json['data']['bondBaseInfo']
        coupon_rate = data.get('parCouponRate')
        
        return {
            'symbol': symbol,
            'maturity_date': data.get('mrtyDate'),
            'coupon_rate': float(coupon_rate) / 100 if coupon_rate and coupon_rate != '---' else 0,
            'frequency': data.get('couponFrqncy', '年'),
            'bond_type': data.get('bondType'),
            'coupon_type': data.get('intrstPayMeth')
        }
    
    def _build_headers(self) -> Dict[str, str]:
        """构建请求头"""
        return {
            "User-Agent": random.choice(self._config.USER_AGENTS),
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": "https://www.chinamoney.com.cn",
            "Referer": "https://www.chinamoney.com.cn/chinese/zqjc/",
            "Connection": "keep-alive",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
        }
    
    def batch_fetch_metadata(self, symbols: List[str], cache: Dict, 
                            normalized_cache: Dict, settlement_dt_str: str,
                            cache_manager: CacheManager) -> int:
        """批量获取元数据"""
        if not symbols:
            return 0
        
        session = self._create_session()
        success_count = 0
        
        with ThreadPoolExecutor(max_workers=self._config.CONCURRENT_THREADS) as executor:
            future_to_symbol = {
                executor.submit(self.fetch_metadata, s, session): s 
                for s in symbols
            }
            
            for future in tqdm(as_completed(future_to_symbol), total=len(symbols), desc="抓取进度"):
                symbol = future_to_symbol[future]
                try:
                    data = future.result()
                    if data:
                        cache[symbol] = data
                        normalized_cache[symbol.replace(" ", "")] = data
                        success_count += 1
                    
                    if success_count % self._config.SAVE_INTERVAL == 0:
                        cache_manager.save_metadata_cache(cache, settlement_dt_str)
                except:
                    pass
        
        cache_manager.save_metadata_cache(cache, settlement_dt_str)
        return success_count


# ==================== 债券计算模块 ====================

class BondCalculator:
    """债券计算器 - 负责所有债券相关计算"""
    
    @staticmethod
    def calculate_remaining_days(maturity_date: Any, settlement_date: Any = None) -> Optional[int]:
        """计算债券剩余天数"""
        try:
            settlement_dt = (datetime.now() if settlement_date is None 
                           else datetime.strptime(settlement_date, '%Y-%m-%d') 
                           if isinstance(settlement_date, str) else settlement_date)
            
            if hasattr(settlement_dt, 'replace'):
                settlement_dt = settlement_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            
            if not maturity_date or maturity_date == '---':
                return None
            
            maturity_dt = (datetime.strptime(maturity_date, '%Y-%m-%d') 
                         if isinstance(maturity_date, str) else maturity_date)
            
            if hasattr(maturity_dt, 'replace'):
                maturity_dt = maturity_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            
            days = (maturity_dt - settlement_dt).days
            return max(0, days) if days > 0 else 0
        except:
            return None
    
    @staticmethod
    def format_tenor(days: Optional[int]) -> str:
        """格式化剩余期限显示"""
        if days is None or days <= 0:
            return ""
        
        years_part = days // 365
        days_part = days % 365
        
        if years_part > 0:
            return f"{years_part}年{days_part}天" if days_part > 0 else f"{years_part}年"
        return f"{days_part}天"
    
    @staticmethod
    def calculate_after_tax_yield(yield_val: Optional[float], bond_type: str) -> Optional[float]:
        """计算税后收益率"""
        if yield_val is None or pd.isna(yield_val):
            return None
        
        is_tax_exempt = (bond_type == '国债' or '地方政府债' in bond_type)
        return yield_val if is_tax_exempt else yield_val * 0.8
    
    @staticmethod
    def day_count_fraction(date1: datetime, date2: datetime, convention: str = 'Act/365') -> float:
        """计算天数分数"""
        days = (date2 - date1).days
        
        if convention == 'Act/Act':
            year1, year2 = date1.year, date2.year
            if year1 == year2:
                days_in_year = 366 if (year1 % 4 == 0 and (year1 % 100 != 0 or year1 % 400 == 0)) else 365
                return days / days_in_year
            return days / 365.25
        elif convention == '30/360':
            d1, d2 = min(30, date1.day), date2.day
            if d1 == 30 and d2 == 31:
                d2 = 30
            return (360 * (date2.year - date1.year) + 30 * (date2.month - date1.month) + (d2 - d1)) / 360
        return days / 365
    
    @staticmethod
    def get_coupon_dates(settlement_date: datetime, maturity_date: datetime, frequency_str: str) -> tuple:
        """生成付息日列表"""
        freq_map = {'年': 1, '半年': 2, '季': 4, '按年付息': 1, '半年付息': 2, '按季付息': 4}
        m = freq_map.get(frequency_str, 1)
        months_step = 12 // m
        
        coupon_dates = []
        current_date = maturity_date
        
        while current_date > settlement_date:
            coupon_dates.append(current_date)
            current_date = current_date - pd.DateOffset(months=months_step)
        
        coupon_dates.sort()
        return coupon_dates, current_date


# ==================== Excel报表模块 ====================

class ExcelReporter:
    """Excel报表生成器"""
    
    def __init__(self, config: Config):
        self._config = config
    
    def generate_report(self, output_file: str, final_df: pd.DataFrame, 
                       header_mapping: Dict, cols_order: List[str]) -> None:
        """生成Excel报表"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._create_combined_sheet(writer, "主页", final_df, header_mapping, cols_order)
            
            # 移除默认的 "Sheet"
            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]

    def _create_combined_sheet(self, writer, sheet_name: str, bonds_df: pd.DataFrame,
                               header_mapping: Dict, cols_order: List[str]) -> None:
        """在同一个sheet中创建三张表格"""
        from openpyxl.utils import get_column_letter
        
        ws = writer.book.create_sheet(sheet_name, 0)
        styles = self._get_styles()
        display_cols = ['债券简称', '税后年收益率', '剩余期限', '到期日']
        
        current_row = 1
        
        # 定义过滤配置
        filters = [
            ("小于6个月到期债券", 0, 180),
            ("小于1年到期债券", 0, 365),
            ("小于3年到期债券", 0, 1095)
        ]
        
        for title, min_days, max_days in filters:
            df = self._filter_bonds(bonds_df, min_days, max_days)
            df = self._apply_filters(df)
            df = self._prepare_df(df, header_mapping, cols_order)
            current_row = self._write_bond_table(ws, current_row, 1, title, df, display_cols, styles)
        
        # 备注
        self._write_notes(ws, current_row, display_cols)
        
        # 列宽设置
        for idx, width in enumerate([20, 12, 15, 12]):
            ws.column_dimensions[get_column_letter(idx + 1)].width = width
        
        # 行高设置
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
    def _filter_bonds(self, df: pd.DataFrame, min_days: int, max_days: int) -> pd.DataFrame:
        """筛选债券"""
        return df[(df['剩余天数'] >= min_days) & (df['剩余天数'] <= max_days)].copy()
    
    def _apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        """应用收益率过滤和成交量筛选"""
        if df.empty:
            return df
        
        max_yield = df['税后年收益率'].max()
        if max_yield is None or pd.isna(max_yield):
            return df
        
        df = df[df['税后年收益率'] >= max_yield * 0.67]
        
        if not df.empty and '交易量' in df.columns:
            df = df.dropna(subset=['交易量'])
            df = df.nlargest(10, '交易量')
        
        return df
    
    def _prepare_df(self, df: pd.DataFrame, header_mapping: Dict, cols_order: List[str]) -> pd.DataFrame:
        """准备DataFrame"""
        if df.empty:
            return df
        
        existing_cols = [c for c in cols_order if c in df.columns]
        df_prepared = df[existing_cols].copy()
        df_prepared.sort_values('税后年收益率', ascending=False, inplace=True)
        df_prepared.rename(columns=header_mapping, inplace=True)
        return df_prepared
    
    def _get_styles(self) -> Dict:
        """获取样式定义"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        return {
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            'title_font': Font(bold=True, size=12),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'yield_fill': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        }
    
    def _write_bond_table(self, ws, start_row: int, start_col: int, title: str,
                          bonds_df: pd.DataFrame, display_cols: List[str], 
                          styles: Dict) -> int:
        """写入债券表格"""
        # 标题
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row, end_column=start_col + len(display_cols) - 1)
        cell = ws.cell(row=start_row, column=start_col)
        cell.value = title
        cell.font = styles['title_font']
        cell.alignment = styles['center']
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        
        # 表头
        header_row = start_row + 1
        for idx, col_name in enumerate(display_cols):
            cell = ws.cell(row=header_row, column=start_col + idx)
            cell.value = col_name
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center']
            cell.border = styles['border']
        
        # 数据
        if bonds_df.empty:
            ws.merge_cells(start_row=header_row + 1, start_column=start_col,
                          end_row=header_row + 1, end_column=start_col + len(display_cols) - 1)
            cell = ws.cell(row=header_row + 1, column=start_col)
            cell.value = "暂无符合条件的债券"
            cell.alignment = styles['center']
            return header_row + 2
        
        for row_idx in range(min(len(bonds_df), 20)):
            excel_row = header_row + 1 + row_idx
            for col_idx, col_name in enumerate(display_cols):
                cell = ws.cell(row=excel_row, column=start_col + col_idx)
                
                if col_name in bonds_df.columns:
                    value = bonds_df.iloc[row_idx][col_name]
                    cell.value = self._format_cell_value(value, col_name)
                else:
                    cell.value = '---'
                
                cell.border = styles['border']
                cell.alignment = styles['center'] if col_idx < 2 else styles['left']
                
                if col_name == '税后年收益率':
                    cell.fill = styles['yield_fill']
        
        return header_row + min(len(bonds_df), 20) + 2
    
    def _format_cell_value(self, value, col_name: str) -> Any:
        """格式化单元格值"""
        if pd.isna(value):
            return '---'
        if isinstance(value, (int, float)):
            if col_name in ['税后年收益率', '票面利率']:
                return round(value, 4)
            if col_name == '成交额(亿元)' and isinstance(value, float) and value.is_integer():
                return int(value)
            return value
        return str(value)
    
    def _write_notes(self, ws, start_row: int, display_cols: List[str]) -> None:
        """写入备注"""
        from openpyxl.styles import Font, Alignment
        
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row + 2, end_column=len(display_cols))
        cell = ws.cell(row=start_row, column=1)
        cell.value = "备注：\n1. 优先按照投资天数需求选择，再根据税后收益率排名获得购买结果。\n2. 所列债券日成交额均大于10亿，流动性有保证。\n3. 推荐购买6个月内的债券，属于无风险的现金等价物。"
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = Font(size=10)


# ==================== 主程序 ====================

class BondAnalysisApp:
    """债券分析应用主类"""
    
    def __init__(self):
        self._config = config
        self._cache_manager = CacheManager(self._config)
        self._data_fetcher = BondDataFetcher(self._config)
        self._calculator = BondCalculator()
        self._reporter = ExcelReporter(self._config)
    
    def run(self) -> None:
        """运行分析流程"""
        # 1. 确定日期和缓存策略
        settlement_dt_str = self._determine_settlement_date()
        
        # 2. 获取成交数据
        deal_df = self._fetch_deal_data(settlement_dt_str)
        if deal_df is None:
            return
        
        deal_df = self._filter_deal_data(deal_df)
        
        # 3. 加载元数据缓存
        cache = self._cache_manager.load_metadata_cache(settlement_dt_str)
        
        # 4. 抓取缺失的元数据
        self._fetch_missing_metadata(deal_df, cache, settlement_dt_str)
        
        # 5. 计算指标
        results = self._calculate_metrics(deal_df, cache, settlement_dt_str)
        
        # 6. 生成报表
        self._generate_report(results, settlement_dt_str)
    
    def _determine_settlement_date(self) -> str:
        """确定结算日期"""
        in_cache_window = self._cache_manager.is_within_cache_window()
        latest_date_str, _ = self._cache_manager.get_latest_cache_date()
        use_cache = (not self._config.ONLINE_MODE) or in_cache_window
        
        if use_cache and latest_date_str:
            if not self._config.ONLINE_MODE:
                print(f"ONLINE_MODE = False，将使用历史最近交易日的缓存数据: {latest_date_str}")
            else:
                print(f"当前处于交易期 (交易日 8:00-20:00)，将使用历史最近交易日的缓存数据: {latest_date_str}")
            return latest_date_str
        elif latest_date_str:
            print(f"当前处于抓取窗口，将优先使用历史缓存数据: {latest_date_str}")
            return latest_date_str
        else:
            print(f"未发现任何历史缓存，将尝试获取最新数据...")
            return datetime.now().strftime("%Y-%m-%d")
    
    def _fetch_deal_data(self, settlement_dt_str: str) -> Optional[pd.DataFrame]:
        """获取成交数据"""
        deal_df = self._data_fetcher.fetch_deal_data(settlement_dt_str, self._cache_manager)
        
        if deal_df is None and not self._config.ONLINE_MODE:
            print("错误: ONLINE_MODE = False 且不存在可用缓存，无法获取数据。")
            return None
        
        return deal_df
    
    def _filter_deal_data(self, deal_df: pd.DataFrame) -> pd.DataFrame:
        """筛选成交数据"""
        initial_count = len(deal_df)
        
        # 筛选国债且非贴现，成交量>=10亿
        mask = deal_df['债券简称'].str.contains('国债') & ~deal_df['债券简称'].str.contains('贴现')
        
        if '交易量' in deal_df.columns:
            mask = mask & (deal_df['交易量'] >= 10)
        
        deal_df = deal_df[mask].copy()
        print(f"统一筛选完成：从 {initial_count} 条过滤至 {len(deal_df)} 条。")
        
        return deal_df
    
    def _fetch_missing_metadata(self, deal_df: pd.DataFrame, cache: Dict, 
                                 settlement_dt_str: str) -> None:
        """抓取缺失的元数据"""
        normalized_cache = {k.replace(" ", ""): v for k, v in cache.items()}
        symbols_to_fetch = [
            s for s in deal_df['债券简称'].unique() 
            if s.replace(" ", "") not in normalized_cache
        ]
        
        if not symbols_to_fetch:
            print("所有成交债券的元数据已在缓存中，跳过抓取。")
            return
        
        if not self._config.ONLINE_MODE:
            print(f"ONLINE_MODE = False，禁止抓取。缺少 {len(symbols_to_fetch)} 个债券的元数据。")
            return
        
        print(f"发现 {len(symbols_to_fetch)} 个新债券缺失元数据，正在抓取...")
        
        session = self._data_fetcher._create_session()
        
        with ThreadPoolExecutor(max_workers=self._config.CONCURRENT_THREADS) as executor:
            future_to_symbol = {
                executor.submit(self._data_fetcher.fetch_metadata, s, session): s 
                for s in symbols_to_fetch
            }
            
            for future in tqdm(as_completed(future_to_symbol), total=len(symbols_to_fetch), desc="抓取进度"):
                symbol = future_to_symbol[future]
                try:
                    data = future.result()
                    if data:
                        cache[symbol] = data
                        normalized_cache[symbol.replace(" ", "")] = data
                    
                    if len(cache) % self._config.SAVE_INTERVAL == 0:
                        self._cache_manager.save_metadata_cache(cache, settlement_dt_str)
                except:
                    pass
        
        self._cache_manager.save_metadata_cache(cache, settlement_dt_str)
        print(f"抓取完成。当前总缓存: {len(cache)} 条。")
    
    def _calculate_metrics(self, deal_df: pd.DataFrame, cache: Dict, 
                           settlement_dt_str: str) -> pd.DataFrame:
        """计算债券指标"""
        print("正在计算剩余期限及久期...")
        
        normalized_cache = {k.replace(" ", ""): v for k, v in cache.items()}
        results = []
        
        session = self._data_fetcher._create_session()
        
        for _, row in tqdm(deal_df.iterrows(), total=len(deal_df), desc="计算进度"):
            symbol = row['债券简称']
            search_key = symbol.replace(" ", "")
            meta = normalized_cache.get(search_key)
            
            # 缓存未命中，尝试实时抓取
            if not meta and self._config.ONLINE_MODE:
                meta = self._data_fetcher.fetch_metadata(symbol, session)
                if meta:
                    cache[symbol] = meta
                    normalized_cache[search_key] = meta
            
            res_row = self._process_row(row, meta, settlement_dt_str)
            results.append(res_row)
        
        self._cache_manager.save_metadata_cache(cache, settlement_dt_str)
        
        final_df = pd.DataFrame(results)
        if final_df.empty:
            print("未发现符合条件的债券数据。")
        
        return final_df
    
    def _process_row(self, row, meta: Optional[Dict], settlement_dt_str: str) -> Dict:
        """处理单行数据"""
        res_row = row.to_dict()
        
        # 格式化交易量
        if '交易量' in res_row and not pd.isna(res_row['交易量']):
            vol = res_row['交易量']
            if isinstance(vol, (float, np.float64)) and vol.is_integer():
                vol = int(vol)
            res_row['交易量'] = vol
        
        if not meta:
            return self._empty_result(res_row)
        
        # 计算收益率
        y_val = row['加权收益率'] if not pd.isna(row['加权收益率']) else row['最新收益率']
        
        # 计算剩余天数
        days = self._calculator.calculate_remaining_days(
            meta['maturity_date'], settlement_dt_str
        )
        
        # 格式化期限
        tenor_display = self._calculator.format_tenor(days)
        
        # 填充元数据
        bond_type = meta.get('bond_type', '')
        after_tax_yield = self._calculator.calculate_after_tax_yield(y_val, bond_type)
        
        res_row.update({
            '到期日': meta['maturity_date'],
            '票面利率': meta['coupon_rate'],
            '付息频率': meta['frequency'],
            '付息方式': meta.get('coupon_type', '---'),
            '剩余期限_格式化': tenor_display,
            '剩余天数': days,
            '债券类型': bond_type,
            '税后年收益率': after_tax_yield
        })
        
        return res_row
    
    def _empty_result(self, res_row: Dict) -> Dict:
        """返回空结果"""
        res_row.update({
            '到期日': None, '票面利率': None, '付息频率': None,
            '付息方式': None, '剩余期限_格式化': None, '剩余天数': None,
            '税后年收益率': None, '债券类型': None
        })
        return res_row
    
    def _generate_report(self, final_df: pd.DataFrame, settlement_dt_str: str) -> None:
        """生成报表"""
        if final_df.empty:
            return
        
        print("正在对债券进行分类并排序...")
        
        output_file = f"{self._config.OUTPUT_FILE_BASE}_{settlement_dt_str}.xlsx"
        
        self._reporter.generate_report(
            output_file, final_df, 
            self._config.HEADER_MAPPING, self._config.COLS_ORDER
        )
        
        print(f"分析完成！结果已保存至: {output_file}")


def main():
    """主函数入口"""
    app = BondAnalysisApp()
    app.run()


if __name__ == "__main__":
    main()
